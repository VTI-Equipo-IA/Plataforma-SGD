# services/exporter.py
"""
Servicio de exportación de planes a Excel.
Genera archivos Excel con formato específico incluyendo estilos y colores.
"""
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from sqlalchemy import select
from extensions.db import db


def export_plans_to_excel(table) -> BytesIO:
    """
    Exporta todos los planes de la base de datos a un archivo Excel
    con el formato especificado.
    
    Args:
        table: Tabla reflejada de SQLAlchemy
        
    Returns:
        BytesIO: Buffer con el contenido del archivo Excel
    """
    # Consultar todos los registros de la base de datos
    stmt = select(table).order_by(table.c.id.asc())
    rows = db.session.execute(stmt).mappings().all()
    
    # Preparar datos para el DataFrame
    data = []
    for row in rows:
        dimension = str(row.get('dimension') or '')
        is_calidad_web = 'calidad web' in dimension.lower()
        is_gobernanza = 'gobernanza' in dimension.lower()
        
        # Función auxiliar para convertir valores None a string vacío
        def safe_str(val):
            return str(val) if val is not None else ''
        
        data.append({
            'Dimensión': safe_str(row.get('dimension')),
            'Instrumento': safe_str(row.get('instrumento')),
            'Subdimension': safe_str(row.get('subdimension')),
            'Brecha': safe_str(row.get('brecha')),
            'Nombre_Iniciativa': safe_str(row.get('iniciativa')),
            'Objetivo de la iniciativa': safe_str(row.get('objetivo_iniciativa')),
            'Indicador_Proceso': safe_str(row.get('indicador_proceso')),
            'Indicador_Impacto': safe_str(row.get('indicador_resultado')),
            'Autor': safe_str(row.get('autor')),
            'Área responsable': '',
            'Costo estimado total': '',
            'Código EVALTIC': '<ID iniciativa EVALTIC, cuando aplique>',
            'N_Actividad_Hito': safe_str(row.get('n_actividad_hito')),
            'Tipo': safe_str(row.get('tipo')),
            'Nombre_Actividad_Hito': safe_str(row.get('descripcion')),
            'Fecha inicio': 'dd-mmm-aaaa',
            'Fecha fin': 'dd-mmm-aaaa',
            'CW-SD_indicador': safe_str(row.get('indicador')) if is_calidad_web else 'n/a',
            'CW-SD_n_preg': safe_str(row.get('n_pregunta')) if is_calidad_web else 'n/a',
            'MGDE_nivel_madurez': safe_str(row.get('nivel_de_madurez')) if is_gobernanza else 'n/a',
        })
    
    # Crear DataFrame
    df = pd.DataFrame(data)
    
    # Crear archivo Excel con openpyxl para aplicar estilos
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Planes PTD"
    
    # Definir estilos
    # Header style - azul oscuro con texto blanco
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Cell alignment
    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Escribir encabezados
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Escribir datos
    for row_idx, row_data in enumerate(df.values, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
            
            # Aplicar color de fondo alterno (gris claro) cada 2 filas
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Ajustar anchos de columna
    column_widths = {
        'A': 35,  # Dimensión
        'B': 30,  # Instrumento
        'C': 30,  # Subdimension
        'D': 50,  # Brecha
        'E': 40,  # Nombre_Iniciativa
        'F': 50,  # Objetivo de la iniciativa
        'G': 35,  # Indicador_Proceso
        'H': 35,  # Indicador_Impacto
        'I': 20,  # Autor
        'J': 25,  # Área responsable
        'K': 20,  # Costo estimado total
        'L': 40,  # Código EVALTIC
        'M': 12,  # N_Actividad_Hito
        'N': 12,  # Tipo
        'O': 50,  # Nombre_Actividad_Hito
        'P': 15,  # Fecha inicio
        'Q': 15,  # Fecha fin
        'R': 25,  # CW-SD_indicador
        'S': 15,  # CW-SD_n_preg
        'T': 20,  # MGDE_nivel_madurez
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Congelar primera fila (encabezados)
    ws.freeze_panes = 'A2'
    
    # Guardar en BytesIO
    wb.save(output)
    output.seek(0)
    
    return output
