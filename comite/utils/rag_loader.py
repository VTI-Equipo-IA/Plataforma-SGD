from __future__ import annotations

import os
from typing import List, Tuple, Callable, Optional

from langchain_openai import OpenAIEmbeddings
from langchain_community.vectorstores import FAISS
from langchain_text_splitters import RecursiveCharacterTextSplitter

from dotenv import load_dotenv
load_dotenv()
api_key = os.getenv("OPENAI_API_KEY") or os.getenv("OPENAI_APIKEY")

# ---- Extractores de texto ----

def _load_txt(path: str) -> str:
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()

def _load_md(path: str) -> str:
    return _load_txt(path)

def _load_pdf(path: str) -> str:
    from PyPDF2 import PdfReader
    reader = PdfReader(path)
    parts = []
    for p in reader.pages:
        t = p.extract_text() or ""
        if t.strip():
            parts.append(t)
    return "\n".join(parts)

def _load_docx(path: str) -> str:
    from docx import Document
    doc = Document(path)
    return "\n".join(p.text for p in doc.paragraphs if p.text)

def _load_xlsx(path: str, sheet_name: Optional[str] = None) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    sh = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    cells = []
    for row in sh.iter_rows(values_only=True):
        for v in row:
            if v is not None:
                cells.append(str(v))
    return "\n".join(cells)

# ---- Utilidades de carga por tipo ----

LOADERS = {
    ".txt": _load_txt,
    ".md": _load_md,
    ".pdf": _load_pdf,
    ".docx": _load_docx,
    ".xlsx": _load_xlsx,
}

def _read_file(path: str) -> str:
    ext = os.path.splitext(path)[1].lower()
    loader = LOADERS.get(ext)
    if not loader:
        return ""
    # xlsx acepta sheet opcional si incluimos "#sheet=Nombre" al final del filename (opcional)
    if ext == ".xlsx" and "#sheet=" in path:
        path_only, sheet = path.split("#sheet=", 1)
        return _load_xlsx(path_only, sheet_name=sheet)
    return loader(path)

def _iter_files(folder: str) -> List[str]:
    out = []
    for root, _, files in os.walk(folder):
        for f in files:
            ext = os.path.splitext(f)[1].lower()
            if ext in LOADERS.keys():
                out.append(os.path.join(root, f))
    return out

# ---- Chunks + Embeddings ----

def _to_documents(texts: List[str], metas: List[dict], chunk_size: int = 1000, chunk_overlap: int = 200):
    splitter = RecursiveCharacterTextSplitter(chunk_size=chunk_size, chunk_overlap=chunk_overlap)
    return splitter.create_documents(texts, metas)

# ---- API pública ----

def build_index_for_folder(
    docs_folder: str,
    index_folder: str,
    emb_model: str = "text-embedding-3-small",
    chunk_size: int = 1000,
    chunk_overlap: int = 200,
) -> str:
    """
    Construye un índice FAISS con todo el contenido legible de `docs_folder`.
    Guarda en `index_folder` y devuelve la ruta del índice.
    """
    files = _iter_files(docs_folder)
    texts: List[str] = []
    metas: List[dict] = []

    for path in files:
        try:
            txt = _read_file(path)
            if not txt or not txt.strip():
                continue
            # metadatos simples
            rel = os.path.relpath(path, docs_folder)
            metas.append({"fuente": rel})
            texts.append(txt)
        except Exception:
            # ignora archivos problemáticos; puedes loggear si lo deseas
            continue

    if not texts:
        os.makedirs(index_folder, exist_ok=True)
        # índice vacío: crea FAISS mínimo para evitar fallos al cargar
        embeddings = OpenAIEmbeddings(model=emb_model)
        FAISS.from_texts([""], embeddings).save_local(index_folder)
        return index_folder

    docs = _to_documents(texts, metas, chunk_size=chunk_size, chunk_overlap=chunk_overlap)
    embeddings = OpenAIEmbeddings(model=emb_model)
    vs = FAISS.from_documents(docs, embeddings)
    os.makedirs(index_folder, exist_ok=True)
    vs.save_local(index_folder)
    return index_folder

def load_retriever(
    index_folder: str,
    emb_model: str = "text-embedding-3-small",
    k: int = 4
) -> Callable[[str], str]:
    """
    Carga un índice FAISS y devuelve una función `retrieve(query) -> str`
    que concatena los k chunks más relevantes en un solo texto.
    """
    embeddings = OpenAIEmbeddings(model=emb_model)
    vs = FAISS.load_local(index_folder, embeddings, allow_dangerous_deserialization=True)

    def retrieve(query: str) -> str:
        docs = vs.similarity_search(query, k=k)
        parts: List[str] = []
        for d in docs:
            src = d.metadata.get("fuente", "?")
            parts.append(f"[{src}] {d.page_content}")
        return "\n\n".join(parts)

    return retrieve

# ---- Ejemplo CLI opcional ----
if __name__ == "__main__":
    # Ejemplo rápido:
    # python utils/rag_loader.py
    DOCS = "documentos/gobernanza/hoja_ruta/Hojas_de_Ruta_MGDE_detallado.docx"          # carpeta con subcarpetas y archivos .docx/.xlsx/.pdf/.txt/.md
    INDEX = "indices/hoja_ruta_index" # destino del índice
    build_index_for_folder(DOCS, INDEX)
    retriever = load_retriever(INDEX)
    print(retriever("accesibilidad WCAG 2.1"))